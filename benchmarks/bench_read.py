"""Benchmark: opensheet_core vs openpyxl for reading XLSX files."""

import os
import sys
import tempfile

import openpyxl
import opensheet_core

from bench_utils import bench_pair, format_bytes, format_time, generate_row


def generate_test_file(path, rows, cols):
    """Generate a test XLSX file using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark"

    ws.append([f"col_{i}" for i in range(cols)])

    for r in range(rows):
        ws.append(generate_row(r, cols))

    wb.save(path)
    return os.path.getsize(path)


def do_openpyxl_read(path):
    """Read all cells with openpyxl (materialize for fair memory comparison)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    _ = len(all_rows)
    wb.close()


def do_opensheet_read(path):
    """Read all cells with opensheet_core."""
    rows = opensheet_core.read_sheet(path)
    _ = len(rows)


def format_speed_relative(ratio):
    if ratio == float("inf"):
        return "inf faster"
    if ratio >= 1:
        return f"{ratio:.1f}x faster"
    return f"{1 / ratio:.1f}x slower"


def format_memory_relative(os_mem, op_mem):
    if os_mem == 0 and op_mem == 0:
        return "no measurable RSS delta"
    if os_mem == 0:
        return "opensheet ~0 RSS delta"
    ratio = op_mem / os_mem
    if ratio >= 1:
        return f"{ratio:.1f}x less RSS delta"
    return f"{1 / ratio:.1f}x more RSS delta"


def format_time_with_std(result):
    """Format min time with mean +/- stddev."""
    base = format_time(result.min_time)
    mean_str = format_time(result.mean_time)
    if result.std_time > 0:
        std_str = format_time(result.std_time)
        return f"{base} (avg {mean_str} +/- {std_str})"
    return base


def run_benchmark(rows, cols, runs=5):
    print(f"\n{'='*60}")
    print(f"Benchmark: {rows:,} rows x {cols} cols ({rows * cols:,} cells)")
    print(f"{'='*60}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    try:
        file_size = generate_test_file(path, rows, cols)
        print(f"File size: {format_bytes(file_size)}")
        print()

        # Warm up
        do_opensheet_read(path)
        do_openpyxl_read(path)

        # Benchmark (interleaved)
        os_result, op_result = bench_pair(
            do_opensheet_read, (path,),
            do_openpyxl_read, (path,),
            runs=runs,
        )

        speedup = op_result.min_time / os_result.min_time if os_result.min_time > 0 else float("inf")
        speed_text = format_speed_relative(speedup)
        mem_text = format_memory_relative(os_result.median_mem, op_result.median_mem)

        print(f"  {'Library':<20} {'Time (min)':<15} {'RSS delta':<15}")
        print(f"  {'-'*50}")
        print(f"  {'opensheet_core':<20} {format_time(os_result.min_time):<15} {format_bytes(os_result.median_mem):<15}")
        print(f"  {'openpyxl':<20} {format_time(op_result.min_time):<15} {format_bytes(op_result.median_mem):<15}")
        print()
        print(f"  Timing:  {format_time_with_std(os_result)}  vs  {format_time_with_std(op_result)}")
        print(f"  Speed:   opensheet_core is {speed_text}")
        print(f"  Memory:  opensheet_core uses {mem_text}")

        return {
            "rows": rows,
            "cols": cols,
            "opensheet_time": os_result.min_time,
            "openpyxl_time": op_result.min_time,
            "opensheet_mem": os_result.median_mem,
            "openpyxl_mem": op_result.median_mem,
            "speedup": speedup,
        }
    finally:
        os.unlink(path)


def main():
    print("OpenSheet Core vs openpyxl — Read Benchmark")
    print(f"opensheet_core {opensheet_core.__version__}")
    print(f"openpyxl {openpyxl.__version__}")
    print(f"Python {sys.version.split()[0]}")
    print(f"Memory: current RSS delta (not high-water mark)")
    print(f"Runs: 5 per config (interleaved)")

    configs = [
        (1_000, 10),
        (10_000, 10),
        (50_000, 10),
        (100_000, 10),
        (10_000, 50),
    ]

    results = []
    for rows, cols in configs:
        result = run_benchmark(rows, cols)
        results.append(result)

    print(f"\n{'='*60}")
    print("Summary")
    print(f"{'='*60}")
    print(f"  {'Config':<20} {'Speed':<16} {'Memory':<24}")
    print(f"  {'-'*56}")
    for r in results:
        config = f"{r['rows']:,} x {r['cols']}"
        print(
            f"  {config:<20} "
            f"{format_speed_relative(r['speedup']):<16} "
            f"{format_memory_relative(r['opensheet_mem'], r['openpyxl_mem']):<24}"
        )


if __name__ == "__main__":
    main()
