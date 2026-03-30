#!/usr/bin/env python3
"""
OpenSheet Core Benchmark Suite

Compare opensheet_core vs openpyxl for reading and writing XLSX files.

Usage:
    python benchmarks/benchmark.py              # default: 100k rows
    python benchmarks/benchmark.py --rows 50000 # custom row count
    python benchmarks/benchmark.py --quick      # fast smoke test (1k rows)
"""

import argparse
import os
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    print("openpyxl is required for benchmarking: pip install openpyxl")
    sys.exit(1)

import opensheet_core

from bench_utils import bench_pair, format_bytes, format_time, generate_row


COLS = 10


# --- Write benchmarks ---

def write_opensheet(path, rows, cols):
    with opensheet_core.XlsxWriter(path) as w:
        w.add_sheet("Benchmark")
        w.write_row([f"col_{i}" for i in range(cols)])
        for r in range(rows):
            w.write_row(generate_row(r, cols))


def write_openpyxl(path, rows, cols):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Benchmark")
    ws.append([f"col_{i}" for i in range(cols)])
    for r in range(rows):
        ws.append(generate_row(r, cols))
    wb.save(path)


# --- Read benchmarks ---

def read_opensheet(path):
    rows = opensheet_core.read_sheet(path)
    _ = len(rows)


def read_openpyxl(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # Materialize all rows to match opensheet's read_sheet() behavior (fair memory comparison)
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    _ = len(all_rows)
    wb.close()


# --- Runner ---

def format_speed_relative(ratio):
    if ratio == float("inf"):
        return "inf faster"
    if ratio >= 1:
        return f"{ratio:.1f}x faster"
    return f"{1 / ratio:.1f}x slower"


def format_memory_relative(os_mem, op_mem):
    if os_mem == 0 and op_mem == 0:
        return "no measurable RSS delta"
    if os_mem == 0:
        return "opensheet ~0 RSS delta"
    ratio = op_mem / os_mem
    if ratio >= 1:
        return f"{ratio:.1f}x less memory"
    return f"{1 / ratio:.1f}x more memory"


def format_time_with_std(result):
    """Format min time with mean +/- stddev."""
    base = format_time(result.min_time)
    mean_str = format_time(result.mean_time)
    if result.std_time > 0:
        std_str = format_time(result.std_time)
        return f"{base} (avg {mean_str} +/- {std_str})"
    return base


def format_mem_with_std(result):
    """Format median memory with mean +/- stddev."""
    base = format_bytes(result.median_mem)
    if result.std_mem > 0:
        std_str = format_bytes(int(result.std_mem))
        return f"{base} (+/- {std_str})"
    return base


def print_comparison(label, os_result, op_result):
    speedup = op_result.min_time / os_result.min_time if os_result.min_time > 0 else float("inf")
    speed_text = format_speed_relative(speedup)
    mem_text = format_memory_relative(os_result.median_mem, op_result.median_mem)

    print(f"\n  {label}")
    print(f"  {'Library':<22} {'Time (min of {})'.format('N'):<14} {'RSS delta (median)':<16}")
    print(f"  {'-'*52}")
    print(f"  {'opensheet_core':<22} {format_time(os_result.min_time):<14} {format_bytes(os_result.median_mem):<16}")
    print(f"  {'openpyxl':<22} {format_time(op_result.min_time):<14} {format_bytes(op_result.median_mem):<16}")
    print()
    print(f"  Timing detail:")
    print(f"    opensheet_core  {format_time_with_std(os_result)}")
    print(f"    openpyxl        {format_time_with_std(op_result)}")
    print(f"  Memory detail:")
    print(f"    opensheet_core  {format_mem_with_std(os_result)}")
    print(f"    openpyxl        {format_mem_with_std(op_result)}")
    print()
    print(f"  -> {speed_text}, {mem_text}")

    return speedup, mem_text


def main():
    parser = argparse.ArgumentParser(description="OpenSheet Core Benchmark Suite")
    parser.add_argument("--rows", type=int, default=100_000, help="Number of rows (default: 100000)")
    parser.add_argument("--cols", type=int, default=COLS, help="Number of columns (default: 10)")
    parser.add_argument("--runs", type=int, default=5, help="Runs per benchmark (default: 5)")
    parser.add_argument("--quick", action="store_true", help="Quick mode: 1000 rows, 1 run")
    args = parser.parse_args()

    if args.quick:
        args.rows = 1_000
        args.runs = 1

    rows, cols, runs = args.rows, args.cols, args.runs

    print("=" * 60)
    print("  OpenSheet Core Benchmark Suite")
    print("=" * 60)
    print(f"  opensheet_core  {opensheet_core.__version__}")
    print(f"  openpyxl        {openpyxl.__version__}")
    print(f"  Python          {sys.version.split()[0]}")
    print(f"  Dataset         {rows:,} rows x {cols} cols ({rows * cols:,} cells)")
    print(f"  Runs            {runs} (interleaved)")
    print(f"  Memory          current RSS delta (not high-water mark)")

    fd, os_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    fd, op_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    try:
        # Warm up (populates OS page cache)
        write_opensheet(os_path, min(rows, 100), cols)
        write_openpyxl(op_path, min(rows, 100), cols)

        # Write benchmark (interleaved)
        os_wr, op_wr = bench_pair(
            write_opensheet, (os_path, rows, cols),
            write_openpyxl, (op_path, rows, cols),
            runs=runs,
        )
        write_speed, write_mem_text = print_comparison("WRITE", os_wr, op_wr)

        os_size = os.path.getsize(os_path)
        op_size = os.path.getsize(op_path)
        print(f"  File sizes: opensheet {format_bytes(os_size)}, openpyxl {format_bytes(op_size)}")

        # Read benchmark (use openpyxl-written file to avoid format skew)
        read_opensheet(op_path)  # warm up
        read_openpyxl(op_path)

        os_rr, op_rr = bench_pair(
            read_opensheet, (op_path,),
            read_openpyxl, (op_path,),
            runs=runs,
        )
        read_speed, read_mem_text = print_comparison("READ", os_rr, op_rr)

        # Summary
        print(f"\n{'=' * 60}")
        print("  SUMMARY")
        print(f"{'=' * 60}")
        print(f"  {'Operation':<10} {'Speed':<18} {'Memory':<24}")
        print(f"  {'-'*50}")
        print(f"  {'Write':<10} {format_speed_relative(write_speed):<18} {write_mem_text:<24}")
        print(f"  {'Read':<10} {format_speed_relative(read_speed):<18} {read_mem_text:<24}")
        print()

    finally:
        for p in (os_path, op_path):
            if os.path.exists(p):
                os.unlink(p)


if __name__ == "__main__":
    main()
